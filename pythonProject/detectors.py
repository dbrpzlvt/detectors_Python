'''
Подготовка данных от детекторов к последующей обработке
'''
import os
import time
import sys
import re

import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
import numpy as np

import tkinter as tk
from tkinter import Radiobutton, Button, Label, messagebox, filedialog, Listbox, Scrollbar
from tkinter.ttk import Combobox
from tkinter.scrolledtext import ScrolledText

from tqdm import tqdm
import check_correct

from logger_setup import setup_logger
logger_FDA, logger_GK = setup_logger()  # логгинг


class Application:

    def __init__(self, parent):
        self.parent = parent
        self.parent.title("Добро пожаловать в приложение!")
        self.parent.geometry("1200x680")

        self.company = tk.StringVar()
        self.gk_pre = tk.StringVar()

        # ставлю значения по умолчанию, чтобы радиокнопки были пустыми при запуске программы
        self.company.set('НИЧЕГО НЕ ВЫБРАНО')
        self.gk_pre.set('НИЧЕГО НЕ ВЫБРАНО')

        # Create a Label to show info for RadioButton
        self.lbl = Label(self.parent, text="От кого\nпоступили данные?", justify=tk.LEFT)
        self.lbl.grid(row=4, column=0, padx=10)

        # Create a RadioButton to ability making choice
        self.fda_rbt = Radiobutton(self.parent, text='ФДА', variable=self.company, value='ФДА',
                                   command=self.visible_gk_options)
        self.gk_rbt = Radiobutton(self.parent, text='ГК', variable=self.company, value='ГК',
                                  command=self.visible_gk_options)
        self.fda_rbt.grid(row=5, column=0, padx=10, sticky=tk.W)
        self.gk_rbt.grid(row=6, column=0, padx=10, sticky=tk.W)

        # Метка и радиокнопки для "ГК"
        self.gk_lbl = Label(self.parent, text="Сделать\nпредобработку?", justify=tk.LEFT)
        self.yes_rbt = Radiobutton(self.parent, text='да', variable=self.gk_pre, value="yes")
        self.no_rbt = Radiobutton(self.parent, text='нет', variable=self.gk_pre, value="no")
        self.yes_rbt.grid_forget()
        self.no_rbt.grid_forget()

        # Главная кнопочка
        self.run_btn = Button(
            self.parent,
            text="Запустить",
            command=lambda: self.selected(self.combobox.get(), self.company.get(), self.gk_pre.get())
        )
        self.run_btn.grid(row=5, column=2, padx=10, pady=10)

        # Create a Label to display an instruction
        self.label = Label(self.parent, text="Обрабатывает файлы замеров, поступивших от ...\n"
                                             "Выберите директорию с исходными файлами или папками с файлами\n"
                                             "Для обратки только что поступивших файлов датчиков можно выбрать папку 'Исходные данные', даже если там находятся подпапки и далее выделить нужные файлы/папки и нажать запустить\n",
                           # "Для обработки файлов, с проверенной структурой можно выбрать папку 'Первичная обработка' (с префиксом PRE_)\n"
                           # "Для отрисовки графиков выбрать папку 'Импортированные данные' (с префиксом IN_)",
                           anchor="w", justify=tk.LEFT)
        self.label.grid(row=0, padx=10, pady=10, columnspan=7, sticky=tk.E)

        # Create a Listbox to display folder contents
        self.folder_contents = Listbox(self.parent, selectmode=tk.EXTENDED, exportselection=False)
        self.folder_contents.grid(row=2, column=0, columnspan=5, sticky=tk.NSEW, padx=10)

        # Create a scrollbar for the Listbox
        self.scrollbar = Scrollbar(self.folder_contents)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.folder_contents.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.configure(command=self.folder_contents.yview)

        # Create a button to browse for a folder
        self.browse_button = Button(self.parent, text="Browse Folder", command=self.browse_folder)
        self.browse_button.grid(row=1, column=0, sticky=tk.W, pady=10, padx=10)

        # Create a Button to select all items in ListBox container
        self.select_button = Button(self.parent, text="Select All", command=self.select_all)
        self.select_button.grid(row=3, column=2, sticky=tk.W, pady=10, padx=10)
        self.editor = ScrolledText(self.parent, wrap='word')
        self.editor.grid(row=2, column=6, padx=(10, 10), sticky=tk.NE)

        # Create a ComboBox to show a menu
        self.combobox = Combobox(self.parent, values=['Предобработка'],
                                 # ['Проверка структуры файлов', 'Предобработка', 'Вставка обработанных данных в шаблон .xlsm', 'Скриншоты'],
                                 width=30, state="readonly")
        self.combobox.grid(row=3, column=0, pady=10, padx=10)
        # self.combobox.bind("<<ComboboxSelected>>", self.DeleteOptions)

        self.cur_year = '2024'
        self.folder_path = None
        self.directory_raw = f'../raw_data/{self.company.get()}/{self.cur_year}/Исходные данные'
        self.directory_pre = f'../raw_data/{self.company.get()}/{self.cur_year}/Первичная обработка'

        # вызвал экземпляр класса Checking с проверкой файлов из скрипта check_correct.py
        self.checking = check_correct.Checking(self.editor, self.parent)

    # Метод для добавления списка содержимого папки в ListBox
    def browse_folder(self):
        self.folder_path = filedialog.askdirectory()  # Open a folder selection dialog
        print(self.folder_path)
        if self.folder_path:
            self.folder_contents.delete(0, tk.END)  # Clear the Listbox
            for item in os.listdir(self.folder_path):
                self.folder_contents.insert(tk.END, item)  # Insert folder contents into Listbox

    # Метод для выбора всего содержимого в ListBox
    def select_all(self):
        self.folder_contents.select_set(0, tk.END)

    def visible_gk_options(self):
        selected_value = self.company.get()
        if selected_value == "ФДА":
            self.gk_lbl.grid_forget()
            self.yes_rbt.grid_forget()
            self.no_rbt.grid_forget()
            self.gk_pre.set('no')
            # print('Выбрано: ' + self.company.get() + ' - ' + self.gk_pre.get())
        elif selected_value == 'ГК':
            self.gk_lbl.grid(row=4, column=1, padx=10)
            self.yes_rbt.grid(row=5, column=1, padx=10, sticky=tk.W)
            self.no_rbt.grid(row=6, column=1, padx=10, sticky=tk.W)
            # print('Выбрано: ' + self.company.get() + ' - ' + self.gk_pre.get())

    def selected(self, event, rbt_company, rbt_gk_pre):
        print('Выбрано: ' + rbt_company + ' - ' + rbt_gk_pre)
        self.directory_raw = f'../raw_data/{self.company.get()}/{self.cur_year}/Исходные данные'
        self.directory_pre = f'../raw_data/{self.company.get()}/{self.cur_year}/Первичная обработка'
        # print(self.directory_raw)
        # print(self.directory_pre)
        # return

        selection = self.combobox.get()
        tmp = [self.folder_contents.get(idx) for idx in
               self.folder_contents.curselection()]  # список папок или файлов из Combobox в окошке
        if not tmp:
            tk.messagebox.showerror(title="ALERT",
                                    message="Не выбран ни один файл или папка из списка.\nВыберите и повторите попытку")
            return
        # print(tmp)
        xlsx_files = []  # список файлов, если в папке "Исходные данные" не файлы, а аще папки
        for i in tmp:  # для каждого файла или папки из Combobox проверяем файл это или папка
            if os.path.isfile(os.path.join(self.folder_path, i)):
                xlsx_files.append(i)  # если файл, то все ок - оставляем список как есть
                # folder_before = os.path.basename(os.path.dirname(os.path.join(self.folder_path, i)))
                # print(folder_before)
                # if folder_before != 'Исходные данные':
                #     xlsx_files.append('/' + folder_before + '/' + i) # если файл внутри подпапки (обычно это название дороги), то добавляем к имени файла еще папку в которой он лежет
                #     print(xlsx_files)
                # elif folder_before in ['Первичная обработка', 'Импортированные данные']:
                #     xlsx_files.append(i)  # если файл, то все ок - оставляем список как есть
            else:
                for file in os.listdir(os.path.join(self.folder_path, i)):
                    xlsx_files.append(os.path.join(i,
                                                   file))  # если папка, не ок - открываем каждую папку и смотрим что там за файл и добавляем в новый список

        # if selection == "Проверка структуры файлов":
        #     self.editor.delete(tk.END)
        #     print(f'Выбор 1.\nПроверяю файлы {chr(10).join(xlsx_files)} на соответствие структуре\n')
        #     self.editor.insert(tk.END, f'\nВыбор 1.\nПроверяю файлы {chr(10).join(xlsx_files)} на соответствие структуре\n')
        #     for filename in tqdm(xlsx_files):
        #         self.editor.insert(tk.END, '\n======' + filename + '======\n')
        #         f = os.path.join(self.folder_path, filename)
        #         if os.path.isfile(f):
        #             self.parent.update_idletasks()
        #             time.sleep(2)
        #             self.structure_check(f)
        # elif selection == "Предобработка":
        if selection == "Предобработка":
            print(f'Выбор 2.\nПредобрабатываю файлы {chr(10).join(xlsx_files)}...\n')
            self.editor.delete(tk.END)
            self.parent.update_idletasks()
            self.editor.insert(tk.END, f'\nВыбор 2.\nПредобрабатываю файлы {chr(10).join(xlsx_files)}...\n')
            for filename in tqdm(xlsx_files):
                self.editor.see(tk.END)
                self.editor.insert(tk.END, f'\n======' + filename + '======\n')
                f = os.path.join(self.folder_path, filename)
                if os.path.isfile(f):
                    self.parent.update_idletasks()
                    time.sleep(2)
                    if rbt_company == 'ФДА':
                        self.editor.insert(tk.END, f'Обработка данных от ФДА...\n'
                                                   f'Проверяю сначала структуру файла {filename}')
                        logger_FDA.info(f'Обработка данных от ФДА...\n'
                                        f'\tПроверяю сначала структуру файла {filename}')
                        self.preprocessing(filename, rbt_company)
                    elif rbt_company == 'ГК':
                        self.editor.insert(tk.END, f'Обработка данных от ГК...\n'
                                                   f'Проверяю сначала структуру файла {filename}')
                        logger_GK.info(f'Обработка данных от ГК...\n'
                                       f'\tПроверяю сначала структуру файла {filename}')
                        self.preprocessing(filename, rbt_company, rbt_gk_pre)
                    else:
                        print('Ничего не выбрано')
                self.editor.insert(tk.END, f'\n{filename} обработан!\n')
                if rbt_company == 'ФДА':
                    logger_FDA.info(f'{filename} обработан!\n')
                elif rbt_company == 'ГК':
                    logger_GK.info(f'{filename} обработан!\n')

    def structure_check(self, file, company_choice):
        # print(f'\n\n====== FILENAME FROM FUNCTION {os.path.join(directory_raw, file)}\n\n')
        # file = 'PRE_км 41+138 а.д М-10 Россия Москва – Санкт-Петербург .xlsx'
        if company_choice == 'ФДА':
            print("Выбор КОМПАНИИ: " + company_choice)
            filename_prefix = re.match('^PRE_.+', str(file))
            if filename_prefix:
                wb = openpyxl.load_workbook(os.path.join(self.directory_pre, file))
            else:
                wb = openpyxl.load_workbook(os.path.join(self.folder_path, file))
            sheet = wb.active

            # 1. Проверка соответствия отчёта заданной форме (с записью итогов проверки в лог-файл)

            if sheet['B3'].value == "Общая интенсивность автомобилей" and sheet['E3'].value == "Легковые (до 6 м)" and \
                    sheet[
                        'H3'].value == "Малые груз. (6-9 м)" and sheet['K3'].value == "Грузовые (9-13 м)" and \
                    sheet['N3'].value == "Груз. большие (13-22 м)" and sheet['Q3'].value == "Автопоезда (22-30 м)" and \
                    sheet[
                        'T3'].value == "Автобусы" and sheet['W3'].value == "Мотоциклы":
                logger_FDA.info(
                        f'{file}: Формат отчёта (Rosautodor_1) - соответствует 7 категориям ТС, Варианту №1 - (шаблон sample_fda_var1.xlsm)')

                # if filename_prefix:
                #     logger_FDA.info(
                #         f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №1 - (шаблон sample_fda_var1.xlsm)')
                #     self.editor.insert(tk.END,
                #                        f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №1\nИспользую шаблон отчета (sample_fda_var1.xlsm)\n')
                #
                # else:
                #     logger_FDA.info(
                #         f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №1 - (шаблон pre_sample_r1.xlsx)')
                #     self.editor.insert(tk.END,
                #                        f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №1\nИспользую шаблон отчета (pre_sample_r1.xlsx)\n')
                return 'Rosautodor_1'
            elif sheet['B3'].value == "Общая интенсивность автомобилей" and sheet[
                'E3'].value == "Легковые (до 4.5 м)" and \
                    sheet[
                        'H3'].value == "Легковые большие (4-6 м)" and sheet[
                'K3'].value == "Малые груз. (6-9 м)" and sheet['N3'].value == "Грузовые (9-13 м)" and \
                    sheet['Q3'].value == "Груз. большие (13-22 м)" and sheet['T3'].value == "Автопоезда (22-30 м)" and \
                    sheet[
                        'W3'].value == "Автобусы" and sheet['Z3'].value == "Мотоциклы":
                logger_FDA.info(
                        f'{file}: Формат отчёта (Rosautodor_2) - соответствует 7 категориям ТС, Варианту №2 - (шаблон pre_sample_r2.xlsx)')

                # if filename_prefix:
                #     logger_FDA.info(
                #         f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №2 - (шаблон sample_fda_var2.xlsm)')
                #     self.editor.insert(tk.END,
                #                        f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №2\nИспользую шаблон отчета (sample_fda_var2.xlsm)\n')
                # else:
                #     logger_FDA.info(
                #         f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №2 - (шаблон pre_sample_r2.xlsx)')
                #     self.editor.insert(tk.END,
                #                        f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №2\nИспользую шаблон отчета (pre_sample_r2.xlsx)\n')

                return 'Rosautodor_2'

            # elif sheet['B6'].value=="За период" and sheet['E7'].value=="легковые автомобили (до 6 м)" and sheet['H7'].value=="микроавтобусы, малые грузовые автомобили (6-9 м)" and sheet['K7'].value=="грузовые автомобили (9-11 м)" and\
            #         sheet['N7'].value=="автобусы (11-13 м)" and sheet['Q7'].value=="грузовые большие автомобили, автопоезда (13-18 м)" and sheet['T7'].value=="длинные автопоезда (> 18 м)":
            #     logger_FDA.info(f'{filename}: Формат отчёта соответствует 6 категориям ТС, Варианту №2')
            # elif sheet['B6'].value == "За период" and sheet['E7'].value == "легковые автомобили (до 6 м)" and sheet['H7'].value == "малые грузовые автомобили до 5 тонн (6-9 м)" and\
            #         sheet['K7'].value == "грузовые автомобили 5-12 тонн (9-11 м)" and sheet['N7'].value == "автобусы (11-13 м)" and sheet['Q7'].value == "грузовые большие автомобили 12-20 тонн (13-22 м)" and\
            #         sheet['T7'].value == "автопоезда более 20 тонн (22-30 м)":
            #     logger_FDA.info(f'{filename}: Формат отчёта соответствует 6 категориям ТС, Варианту №3')
            #     editor.insert(tk.END, f'{file}: Формат отчёта соответствует 6 категориям ТС, Варианту №3')
            else:
                logger_FDA.info(f'{file}: Формат отчёта не совпадает')
                self.editor.insert(tk.END, f'{file}: Формат отчёта не совпадает')
                return None
        elif company_choice == 'ГК':
            filename_prefix = re.match('^PRE_.+', str(file))
            if filename_prefix:
                wb = openpyxl.load_workbook(os.path.join(self.directory_pre, file))
            else:
                wb = openpyxl.load_workbook(os.path.join(self.folder_path, file))
            sheet = wb.active

            if sheet['B6'].value == "За период" and sheet['E7'].value == "легковые" and sheet[
                'H7'].value == "микроавтобусы, малые " \
                               "грузовики" and sheet[
                'K7'].value == "одиночные АТС, автобусы" and \
                    sheet['N7'].value == "автопоезда до 13 м" and sheet['Q7'].value == "автопоезда 13..18  м" and sheet[
                'T7'].value == "длинные автопоезда свыше 18 м":
                logger_GK.info(f'{file}: Формат отчёта (Autodor_1) - соответствует 6 категориям ТС, Варианту №1')
                return 'Autodor_1'

            elif sheet['B6'].value == "За период" and sheet['E7'].value == "легковые автомобили (до 6 м)" and sheet[
                'H7'].value == "микроавтобусы, малые грузовые автомобили (6-9 м)" and sheet[
                'K7'].value == "грузовые автомобили (9-11 м)" and \
                    sheet['N7'].value == "автобусы (11-13 м)" and sheet[
                'Q7'].value == "грузовые большие автомобили, автопоезда (13-18 м)" and sheet[
                'T7'].value == "длинные автопоезда (> 18 м)":
                logger_GK.info(f'{file}: Формат отчёта (Autodor_2) - соответствует 6 категориям ТС, Варианту №2')
                return 'Autodor_2'

            elif sheet['B6'].value == "За период" and sheet['E7'].value == "легковые автомобили (до 6 м)" and sheet[
                'H7'].value == "малые грузовые автомобили до 5 тонн (6-9 м)" and \
                    sheet['K7'].value == "грузовые автомобили 5-12 тонн (9-11 м)" and sheet[
                'N7'].value == "автобусы (11-13 м)" and sheet[
                'Q7'].value == "грузовые большие автомобили 12-20 тонн (13-22 м)" and \
                    sheet['T7'].value == "автопоезда более 20 тонн (22-30 м)":
                logger_GK.info(f'{file}: Формат отчёта (Autodor_3) - соответствует 6 категориям ТС, Варианту №3')
                return 'Autodor_3'

            else:
                logger_GK.info(f'{file}: Формат отчёта не совпадает')
                return None

    def preprocessing(self, file, company_choice, gk_pre_choice=None):

        type_file = self.structure_check(file, company_choice)

        if type_file:
            print(f'Все ок! Тип файла {file} замеров известен - {type_file}')
        else:
            # tk.messagebox.showerror(title="ALERT",
            #                         message=f"Тип файла не был возвращен или не существует в файле.\n"
            #                                 f"Проверьте файл {file} на соотствение структуре и повторите попытку, либо удалите этот файл из обработки")
            print(f"Тип файла не был возвращен или не существует в файле.\n"
                  f"Проверьте файл {file} на соотствение структуре и повторите попытку, либо удалите этот файл из обработки\n")
            return

        if company_choice == 'ФДА':
            self.do_FDA(file, type_file, company_choice)
        elif company_choice == 'ГК':
            self.do_GK(file, type_file, company_choice, gk_pre_choice)

    def do_GK(self, file, type_file, company_choice, gk_pre_choice):

        print(f"УРААА!!! МЫ ВНУТРИ ГК АВТОДОР\n"
              f"Обрабатываю файл {file}")
        if gk_pre_choice == 'yes':
            wb = openpyxl.load_workbook(os.path.join(self.folder_path, file))
            sheet = wb.active

            print('2. Снимаю объединение ячеек в строках ниже восьмой')
            # 2. Снимаю объединение ячеек в строках ниже восьмой
            merged_cells = list(map(str, sheet.merged_cells.ranges))
            #    print(merged_cells)
            # Вначале исключаю из рассмотрения элементы в строках 1-8
            for item in merged_cells:
                if len(re.findall('\D[1-8]:', item)) > 0:
                    merged_cells.remove(item)
            # Решение ниже заимствовано с https://qna.habr.com/q/1241978
            # Разъединяю объединенные ячейки и дублирую запись
            #    print(merged_cells)
            for item in merged_cells:
                sheet.unmerge_cells(item)
                merged_cells_range = item.split(":")
                if merged_cells_range[0][0] == merged_cells_range[1][0]:
                    letter = item.split(":").pop(0)[0]  # Символ столбца диапазона
                    # Насколько я понимаю, решение пригодно только для не очень широких таблиц, поскольку под символ столбца отводится строго один символ
                    start = int(item.split(":").pop(0)[1:])  # Начало диапазона
                    end = int(item.split(":").pop()[1:])  # Конец диапазона
                    copy_cell = sheet[(letter + str(start))].value
                    for n in range(start, end + 1):
                        cell = letter + str(n)
                        sheet[cell].value = copy_cell

            print('3. В столбце А преобразовываю дату и время в нужный формат')
            #    3. В столбце А преобразовываю дату и время в нужный формат
            columnA = sheet['A']
            date = str("0")

            # Для каждой строки выполняю проверки: если в содержимом ячейки есть "2022" или "2023" - это отсылка к дате. Тогда выполняю обновление значения локальной переменной "дата" (date)
            # Если в записи ячейки есть цифры и между ними тире, значит, в ней указан диапазон времени.
            # Ячейки с диапазоном времени дополняются актуальной записью о дате и преобразовываются в необходимый формат (дополняются минутами и секундами)
            for row in columnA:
                if re.search("2023", str(row.value)) or re.search("2024", str(row.value)) or re.search("2025", str(row.value)):
                    date = str(row.value)
                if re.search('\d.+-\d.+', str(row.value)):  # Определяю наличие фрагмента  "[цифра(ы)]-[цифра(ы)]"
                    #        row.value=str(date + " " + str(row.value[0:2]) + ":00:00-" + str(row.value[0:2]) + ":59:59")
                    row.value = str(date + " " + str(row.value[0:2]) + ":59:59")

            sheet.column_dimensions[
                'A'].width = 23  # Изменяю ширину столбца после преобразований для более комфортного представления данных

            print('4. Убираю пустые строки, где указана только дата')
            # 4. Убираю пустые строки, где указана только дата
            # Проверку начинаю с 8 строки (после всех заголовков). Иду от большего к меньшему, чтобы не съезжала нумерация строк.
            # Учитываю, что строки в таблице (в библиотеке openpyxl) нумеруются от единицы, а элементы в массиве - от нуля

            for row in range(len(columnA) - 1, 7, -1):
                if re.search("59:59", str(columnA[row].value)) == None:
                    sheet.delete_rows(row + 1, 1)
                    # print("Удаляю строку" + str(row+1))

            # sheet2 = wb.create_sheet('My Sheet New')
            #
            # for row in sheet.iter_rows(min_row=9, values_only=True):  # sheet.values:  # вот тут необходимо начать с 9 строки
            #     if all(x is None for x in row[1:-1]):
            #         continue
            #     sheet2.append(row)
            #
            # from openpyxl.utils import get_column_letter, column_index_from_string
            #
            # sheet.cell(row=9, column=1) = sheet2.values
            # sheet.range(str('A9'+':'+str(get_column_letter(sheet.max_column)))+str(sheet.max_row)).value = sheet2.values
            #
            # del sheet2
            # # sheet2.title = "Page 2"
            # wb.save(os.path.join('C:\\Users\\RomanBevz\\Documents\\detectors_Python\\raw_data\\ГК\\2024\\filename.xlsx'))
            #
            # iter_rows_generator = sheet.iter_rows(values_only=True)
            # for row in iter_rows_generator:  # reversed([*iter_rows]):
            #     print(all(x is None for x in row[1:-1]))
            #
            #     print(any(row), all(row))
            #     # all() return False if all of the row value is None
            #     # if not any(cell.value for cell in row):
            #     #     print(row)

            # Добавляю слева пустой столбец
            # sheet.insert_cols(0)

            # Блок сохранения результатов в новый файл

            print('5. Формирую название нового файла')
            # Формирую название нового файла
            road_name_cell = sheet['A3'].value
            road_name = re.search('[^,].+,', str(road_name_cell))[
                0]  # Выбираю фрагмент текстовой записи от её начала до символа запятой
            road_name = road_name[0:len(road_name) - 1]  # Удаляю последний символ (запятую)
            road_name = road_name.replace('/', '.')  # Слеш заменяю на точку
            road_name = road_name.replace('"',
                                          '')  # Удаляю символы кавычек, вопросительный и восклицательный знаки, символ звёздочки, которые могут мешать сохранению файла
            road_name = road_name.replace('?', '')
            road_name = road_name.replace('!', '')
            road_name = road_name.replace('*', '')

            #    piketage = re.search('\d+', str(filename)[10:])[0]
            #    piketage = piketage[3:]
            #    filename_out=str(filename)[0]+re.search('\d+', str(filename)[2:7])[0]+'_km'+piketage+'.xlsx'
            filename_out = 'PRE_' + str(road_name) + '.xlsx'
            sheet.title = 'Исходные данные'  # Переименовываю лист

            # directory2 = 'Первичная обработка'

            print(f'Сохраняю {filename_out}')
            try:
                wb_out = openpyxl.load_workbook(os.path.join(self.directory_pre, filename_out))
                sheet2 = wb_out.active
                # Создаю кортеж с новыми данными (по новому кварталу) для последующего дополнения существующей таблицы
                sheet_cells = []
                for rows in sheet.iter_rows():
                    row_cells = []
                    for cell in rows:
                        row_cells.append(cell.value)
                    sheet_cells.append(tuple(row_cells))
                # Убираю заголовки из таблицы-дополнения
                sheet_cells = sheet_cells[8:]
                #        print (sheet_cells)

                for row in sheet_cells:
                    sheet2.append(row)

                wb_out.save(os.path.join(self.directory_pre, filename_out))
            except:
                wb.save(os.path.join(self.directory_pre, filename_out))

        elif gk_pre_choice == 'no':

            # file = 'PRE_А-105 подъезд к а.п Домодедово км 43+100 лево (перенесен км 50).xlsx'
            # file = 'PRE_А-105 подъезд к а.п Домодедово км 43+100 прямое.xlsx'
            # file = 'PRE_А-107 ММК Минско-Можайское шоссе км 0+870.xlsx'
            # file = 'PRE_М-1, км  44+000 (перенесен на км 52).xlsx'
            # file = 'PRE_М-1, км 387+600 (перенесен на км 395)'
            # file = 'PRE_М-3 Украина км 249.xlsx'
            # file = 'PRE_М-3 Украина км 176 (демонтирован).xlsx'
            # file = 'PRE_М-1, км  19 (перенесен на км 39).xlsx'
            # создаю ID детектора
            try:
                if bool(re.search('перенесен', file)):
                    if bool(re.search('прямое|обратное|лево|право', file)):
                        detector_id = re.search(r'[A-Z|А-Я]-\d+', file).group(0).lower() \
                                          .replace('-', '') + '_km' + re.search(r'\d+(?=\+)', file).group(0) + \
                                      '_moved_to_' + re.search(r'\d+(?=\))', file).group(0) + '_' + re.search(r'прямое|обратное|лево|право', file).group(0)
                    else:
                        detector_id = re.search(r'[A-Z|А-Я]-\d+', file).group(0).lower() \
                                          .replace('-', '') + '_km' + re.search(r'\d+(?=\+)', file).group(0) + \
                                      '_moved_to_' + re.search(r'\d+(?=\))', file).group(0)
                elif bool(re.search('прямое|обратное|лево|право', file)):
                    detector_id = re.search(r'[A-Z|А-Я]-\d+', file).group(0).lower() \
                                      .replace('-', '') + '_km' + re.search(r'\d+(?=\+)', file).group(0) + '_' + \
                                  re.search(r'прямое|обратное|лево|право', file).group(0)
                elif bool(re.search('объединенный', file)):
                    detector_id = file
                elif bool(re.search('демонтирован', file)):
                    detector_id = file
                elif bool(re.search(' в | из ', file)):
                    detector_id = file
                elif bool(re.search('альтернатива', file)):
                    detector_id = re.search(r'[A-Z|А-Я]-\d+', file).group(0).lower() \
                                      .replace('-', '') + '_km' + re.search(r'\d+(?=\+)', file).group(0) + '_' + \
                                  re.search(r'альтернатива', file).group(0)
                else:
                    detector_id = re.search(r'[A-Z|А-Я]-\d+', file).group(0).lower() \
                                      .replace('-', '') + '_km' + re.search(r'\d+(?=\+)', file).group(0)
            except:
                detector_id = file

            # читаю предобработанный файл
            df = self.checking.open_and_read_file(self.directory_pre, file, type_file)
            # Делаю длинный формат и считаю корректность данных
            df_total_long = self.checking.make_long(df, company_choice, self.cur_year, file, type_file)
            df_total_long_reserved = df_total_long.copy()
            df_main_clear, basic_stats_SSID, basic_stats_intensivnosti = self.checking.fill_gaps_and_remove_outliers(
                df_total_long, detector_id, type_file)
            # Рисую графики
            self.checking.plot_graphs(self.company.get(), df_main_clear, df_total_long, self.cur_year, file, freq='d')
            # Сохраняю статистику
            bs_SSID = []
            bs_intensivnosti = []
            bs_SSID.append(basic_stats_SSID)
            bs_intensivnosti.append(basic_stats_intensivnosti)

            # bs_SSID_full = pd.DataFrame()
            # bs_intensivnosti_full = pd.DataFrame()
            bs_SSID_full = pd.concat(bs_SSID)
            bs_intensivnosti_full = pd.concat(bs_intensivnosti)

            if os.path.isfile(f'../out/{company_choice}/basic_stats_SSID.xlsx'):  # if file already exists append to existing file
                wb_basic_stats = openpyxl.load_workbook(
                    f'../out/{company_choice}/basic_stats_SSID.xlsx')  # load workbook if already exists
                for direction in ['Итого', 'Прямое', 'Обратное']:
                    # dir = 'Прямое'
                    # workbook = openpyxl.load_workbook(file)
                    # workbook = xlsxwriter.Workbook(workbook_file)
                    try:
                        sheet_basic_stats = wb_basic_stats[direction]  # declare the active sheet
                        bs_SSID_to_excel = bs_SSID_full.reset_index() \
                            .rename(columns={'level_0': 'detector_id'}) \
                            .assign(idx=pd.factorize(bs_SSID_full['index'])[0]) \
                            .pivot(index='detector_id', columns=['idx', 'index'], values=direction) \
                            .droplevel('idx', axis=1).reset_index()
                        bs_SSID_to_excel['день максимума'] = bs_SSID_to_excel['день максимума'].astype(str)
                        # bs_SSID_to_excel.dtypes
                        # append the dataframe results to the current excel file
                        for row in dataframe_to_rows(bs_SSID_to_excel, header=False, index=False):
                            sheet_basic_stats.append(row)
                    except KeyError:  # если листа с таким названием нет
                        wb_basic_stats.create_sheet(direction)
                        sheet_basic_stats = wb_basic_stats[direction]  # declare the active sheet
                        bs_SSID_to_excel = bs_SSID_full.reset_index() \
                            .rename(columns={'level_0': 'detector_id'}) \
                            .assign(idx=pd.factorize(bs_SSID_full['index'])[0]) \
                            .pivot(index='detector_id', columns=['idx', 'index'], values=direction) \
                            .droplevel('idx', axis=1).reset_index()
                        bs_SSID_to_excel['день максимума'] = bs_SSID_to_excel['день максимума'].astype(str)
                        # bs_SSID_to_excel.dtypes
                        # append the dataframe results to the current excel file
                        for row in dataframe_to_rows(bs_SSID_to_excel, header=True, index=False):
                            sheet_basic_stats.append(row)
                wb_basic_stats.save(f'../out/{company_choice}/basic_stats_SSID.xlsx')  # save workbook
                wb_basic_stats.close()  # close workbook
            else:  # create the excel file if doesn't already exist
                for direction in ['Итого', 'Прямое', 'Обратное']:
                    bs_SSID_to_excel = bs_SSID_full.reset_index() \
                        .rename(columns={'level_0': 'detector_id'}) \
                        .assign(idx=pd.factorize(bs_SSID_full['index'])[0]) \
                        .pivot(index='detector_id', columns=['idx', 'index'], values=direction) \
                        .droplevel('idx', axis=1).reset_index()
                    bs_SSID_to_excel['день максимума'] = bs_SSID_to_excel['день максимума'].astype(str)
                    with pd.ExcelWriter(path=f'../out/{company_choice}/basic_stats_SSID.xlsx', engine='openpyxl') as writer:
                        bs_SSID_to_excel.to_excel(writer, index=False, header=True, sheet_name=direction)

            month_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September',
                          'October', 'November', 'December']
            weekdays_list = ["Monday", "Tuesday", "Wednesday", "Thursday",
                             "Friday", "Saturday", "Sunday"]
            hours_list = ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10',
                          '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21',
                          '22', '23']

            if os.path.isfile(
                    f'../out/{company_choice}/basic_stats_intensivnosti.xlsx'):  # if file already exists append to existing file
                wb_basic_stats = openpyxl.load_workbook(
                    f'../out/{company_choice}/basic_stats_intensivnosti.xlsx')  # load workbook if already exists
                for direction in ['Итого', 'Прямое', 'Обратное']:
                    # dir = 'Прямое'
                    # workbook = openpyxl.load_workbook(file)
                    # workbook = xlsxwriter.Workbook(workbook_file)
                    try:
                        sheet_basic_stats = wb_basic_stats[direction]  # declare the active sheet
                        bs_intensivnosti_full = bs_intensivnosti_full.reset_index().rename(
                            columns={'level_0': 'detector_id'})
                        bs_intensivnosti_full['index'] = pd.Categorical(bs_intensivnosti_full['index'],
                                                                        categories=month_list + weekdays_list + hours_list,
                                                                        ordered=True)
                        bs_intensivnosti_full.sort_values(by='index', inplace=True)
                        bs_intensivnosti_full = bs_intensivnosti_full.reset_index(drop=True).set_index(
                            ['detector_id', 'index'])
                        bs_intensivnosti_to_excel = bs_intensivnosti_full.iloc[:,
                                                    bs_intensivnosti_full.columns.get_level_values(0) == direction]
                        bs_intensivnosti_to_excel.columns = bs_intensivnosti_to_excel.columns.droplevel(0)
                        bs_intensivnosti_to_excel = bs_intensivnosti_to_excel.reset_index(drop=False)
                        # append the dataframe results to the current excel file
                        for row in dataframe_to_rows(bs_intensivnosti_to_excel, header=False, index=False):
                            sheet_basic_stats.append(row)
                    except KeyError:  # если листа с таким названием нет
                        wb_basic_stats.create_sheet(direction)
                        sheet_basic_stats = wb_basic_stats[direction]  # declare the active sheet
                        bs_intensivnosti_full = bs_intensivnosti_full.reset_index().rename(
                            columns={'level_0': 'detector_id'})
                        bs_intensivnosti_full['index'] = pd.Categorical(bs_intensivnosti_full['index'],
                                                                        categories=month_list + weekdays_list + hours_list,
                                                                        ordered=True)
                        bs_intensivnosti_full.sort_values(by='index', inplace=True)
                        bs_intensivnosti_full = bs_intensivnosti_full.reset_index(drop=True).set_index(
                            ['detector_id', 'index'])
                        bs_intensivnosti_to_excel = bs_intensivnosti_full.iloc[:,
                                                    bs_intensivnosti_full.columns.get_level_values(0) == direction]
                        bs_intensivnosti_to_excel.columns = bs_intensivnosti_to_excel.columns.droplevel(0)
                        bs_intensivnosti_to_excel = bs_intensivnosti_to_excel.reset_index(drop=False)
                        # append the dataframe results to the current excel file
                        for row in dataframe_to_rows(bs_intensivnosti_to_excel, header=True, index=False):
                            sheet_basic_stats.append(row)
                wb_basic_stats.save(f'../out/{company_choice}/basic_stats_intensivnosti.xlsx')  # save workbook
                wb_basic_stats.close()  # close workbook
            else:  # create the excel file if doesn't already exist
                for direction in ['Итого', 'Прямое', 'Обратное']:
                    bs_intensivnosti_full = bs_intensivnosti_full.reset_index().rename(
                        columns={'level_0': 'detector_id'})
                    bs_intensivnosti_full['index'] = pd.Categorical(bs_intensivnosti_full['index'],
                                                                    categories=month_list + weekdays_list + hours_list,
                                                                    ordered=True)
                    bs_intensivnosti_full.sort_values(by='index', inplace=True)
                    bs_intensivnosti_full = bs_intensivnosti_full.reset_index(drop=True).set_index(
                        ['detector_id', 'index'])
                    bs_intensivnosti_to_excel = bs_intensivnosti_full.iloc[:,
                                                bs_intensivnosti_full.columns.get_level_values(0) == direction]
                    bs_intensivnosti_to_excel.columns = bs_intensivnosti_to_excel.columns.droplevel(0)
                    bs_intensivnosti_to_excel = bs_intensivnosti_to_excel.reset_index(drop=False)
                    # .assign(idx=pd.factorize(bs_intensivnosti_full['index'])[0]) \
                    # .pivot(index='detector_id', columns=['idx', 'index'], values=direction) \
                    # .droplevel('idx', axis=1).reset_index()
                    with pd.ExcelWriter(path=f'../out/{company_choice}/basic_stats_intensivnosti.xlsx', engine='openpyxl') as writer:
                        bs_intensivnosti_to_excel.to_excel(writer, index=False, header=True, sheet_name=direction)

            # wb_data.app.quit()

    def do_FDA(self, file, type_file, company_choice):

        wb_data = xw.Book(os.path.join(self.folder_path, file))

        wb = openpyxl.load_workbook(os.path.join(self.folder_path, file))  # открываем файл с данными
        sheet = wb.active  # активируем лист
        columnA = sheet['A']  # запоминаем колонку А

        dataset_begins = []  # список ячеек, где хранятся текстовые данные для аккуратных наименований

        for cell in columnA:
            if (re.search("км", str(cell.value)) or re.search("km", str(cell.value))) and cell.row > 4:
                dataset_begins = dataset_begins + [cell.row]
        dataset_begins = dataset_begins + [len(columnA) + 1]
        # Копирую данные о названии дороги и местоположении детектора, категориях ТС
        for i in range(0, len(dataset_begins) - 1):
            wb_sample = None
            if type_file == 'Rosautodor_1':
                wb_sample = xw.Book('../pre_sample_r1.xlsx')  # pre_sample_r1.xlsx
            elif type_file == 'Rosautodor_2':
                wb_sample = xw.Book('../pre_sample_r2.xlsx')  # pre_sample_r2.xlsx

            road = wb_data.sheets[0]['A5'].value
            wb_sample.sheets['Исходные данные']['A5'].value = road

            place = wb_data.sheets[0]['A' + str(dataset_begins[i])].value
            wb_sample.sheets['Исходные данные']['A6'].value = place

            place = place.replace('/', '.')  # Слеш заменяю на точку
            place = place.replace('"',
                                  '')  # Удаляю символы кавычек, вопросительный и восклицательный знаки, символ звёздочки, которые могут мешать сохранению файла
            place = place.replace('?', '')
            place = place.replace('!', '')
            place = place.replace('*', '')
            place = place.replace('\n', '')  # Убираю символы переноса строки
            place = place.replace('\t', '')  # Убираю символы табуляции

            # Копирую исходные данные об интенсивности движения
            if type_file == 'Rosautodor_1':
                # my_values = self.wb_data.sheets[0].range('A7:Y' + str(last_row)).options(ndim=2).value
                my_values = wb_data.sheets[0].range(
                    str('A' + str(dataset_begins[i] + 1) + ':Y' + str(dataset_begins[i + 1] - 1))).options(
                    ndim=2).value

            elif type_file == 'Rosautodor_2':
                # my_values = self.wb_data.sheets[0].range('A7:AB' + str(last_row)).options(ndim=2).value
                my_values = wb_data.sheets[0].range(
                    str('A' + str(dataset_begins[i] + 1) + ':AB' + str(dataset_begins[i + 1] - 1))).options(
                    ndim=2).value

            wb_sample.sheets['Исходные данные'].range('A7').value = my_values

            # # Копирую заголовок с описанием типов транспортных средств
            # car_category=wb_data.sheets['Исходные данные'].range('E7:AB7').options(ndim=2).value
            # wb_sample.sheets['Исходные данные'].range('E1:AB1').value = car_category

            # Сохранение под новым названием в специальной папке
            # wb_sample.save(os.path.join('Первичная обработка', 'PRE_' + place + '.xlsx'))
            new_file_name = 'PRE_' + place + '.xlsx'
            self.editor.insert(tk.END, f'\n===== Обрабатываю файл {new_file_name}... =====\n\n')
            wb_sample.save(os.path.join(self.directory_pre, new_file_name))
            wb_sample.close()

            # читаю предобработанный файл
            df = self.checking.open_and_read_file(self.directory_pre, new_file_name, type_file)

            # создаю ID детектора
            if 'ММЗ' in new_file_name:
                detector_id = re.search(r'[A-Z|А-Я]-\d+', new_file_name).group(0).lower() \
                                  .replace('-', '') + '_km' + re.search(r'\d+(?=\+)', new_file_name).group(0) + '_mv'
            else:
                detector_id = re.search(r'[A-Z|А-Я]-\d+', new_file_name).group(0).lower() \
                                  .replace('-', '') + '_km' + re.search(r'\d+(?=\+)', new_file_name).group(0)

            df_total_long = self.checking.make_long(df, self.company.get(), self.cur_year, new_file_name, type_file)
            df_main_clear, basic_stats_SSID, basic_stats_intensivnosti = self.checking.fill_gaps_and_remove_outliers(
                df_total_long, detector_id, type_file)

            # созраняю статистику
            bs_SSID = []
            bs_intensivnosti = []
            bs_SSID.append(basic_stats_SSID)
            bs_intensivnosti.append(basic_stats_intensivnosti)

            bs_SSID_full = pd.concat(bs_SSID)
            bs_intensivnosti_full = pd.concat(bs_intensivnosti)

            if os.path.isfile(f'../out/{company_choice}/basic_stats_SSID.xlsx'):  # if file already exists append to existing file
                wb_basic_stats = openpyxl.load_workbook(
                    f'../out/{company_choice}/basic_stats_SSID.xlsx')  # load workbook if already exists
                for direction in ['Итого', 'Прямое', 'Обратное']:
                    # dir = 'Прямое'
                    # workbook = openpyxl.load_workbook(file)
                    # workbook = xlsxwriter.Workbook(workbook_file)
                    try:
                        sheet_basic_stats = wb_basic_stats[direction]  # declare the active sheet
                        bs_SSID_to_excel = bs_SSID_full.reset_index() \
                            .rename(columns={'level_0': 'detector_id'}) \
                            .assign(idx=pd.factorize(bs_SSID_full['index'])[0]) \
                            .pivot(index='detector_id', columns=['idx', 'index'], values=direction) \
                            .droplevel('idx', axis=1).reset_index()
                        bs_SSID_to_excel['день максимума'] = bs_SSID_to_excel['день максимума'].astype(str)
                        # bs_SSID_to_excel.dtypes
                        # append the dataframe results to the current excel file
                        for row in dataframe_to_rows(bs_SSID_to_excel, header=False, index=False):
                            sheet_basic_stats.append(row)
                    except KeyError:  # если листа с таким названием нет
                        wb_basic_stats.create_sheet(direction)
                        sheet_basic_stats = wb_basic_stats[direction]  # declare the active sheet
                        bs_SSID_to_excel = bs_SSID_full.reset_index() \
                            .rename(columns={'level_0': 'detector_id'}) \
                            .assign(idx=pd.factorize(bs_SSID_full['index'])[0]) \
                            .pivot(index='detector_id', columns=['idx', 'index'], values=direction) \
                            .droplevel('idx', axis=1).reset_index()
                        bs_SSID_to_excel['день максимума'] = bs_SSID_to_excel['день максимума'].astype(str)
                        # bs_SSID_to_excel.dtypes
                        # append the dataframe results to the current excel file
                        for row in dataframe_to_rows(bs_SSID_to_excel, header=True, index=False):
                            sheet_basic_stats.append(row)
                wb_basic_stats.save(f'../out/{company_choice}/basic_stats_SSID.xlsx')  # save workbook
                wb_basic_stats.close()  # close workbook
            else:  # create the excel file if doesn't already exist
                for direction in ['Итого', 'Прямое', 'Обратное']:
                    bs_SSID_to_excel = bs_SSID_full.reset_index() \
                        .rename(columns={'level_0': 'detector_id'}) \
                        .assign(idx=pd.factorize(bs_SSID_full['index'])[0]) \
                        .pivot(index='detector_id', columns=['idx', 'index'], values=direction) \
                        .droplevel('idx', axis=1).reset_index()
                    bs_SSID_to_excel['день максимума'] = bs_SSID_to_excel['день максимума'].astype(str)
                    with pd.ExcelWriter(path=f'../out/{company_choice}/basic_stats_SSID.xlsx', engine='openpyxl') as writer:
                        bs_SSID_to_excel.to_excel(writer, index=False, header=True, sheet_name=direction)

            month_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September',
                          'October', 'November', 'December']
            weekdays_list = ["Monday", "Tuesday", "Wednesday", "Thursday",
                             "Friday", "Saturday", "Sunday"]
            hours_list = ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10',
                          '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21',
                          '22', '23']

            if os.path.isfile(
                    f'../out/{company_choice}/basic_stats_intensivnosti.xlsx'):  # if file already exists append to existing file
                wb_basic_stats = openpyxl.load_workbook(
                    f'../out/{company_choice}/basic_stats_intensivnosti.xlsx')  # load workbook if already exists
                for direction in ['Итого', 'Прямое', 'Обратное']:
                    # dir = 'Прямое'
                    # workbook = openpyxl.load_workbook(file)
                    # workbook = xlsxwriter.Workbook(workbook_file)
                    try:
                        sheet_basic_stats = wb_basic_stats[direction]  # declare the active sheet
                        bs_intensivnosti_full = bs_intensivnosti_full.reset_index().rename(
                            columns={'level_0': 'detector_id'})
                        bs_intensivnosti_full['index'] = pd.Categorical(bs_intensivnosti_full['index'],
                                                                        categories=month_list + weekdays_list + hours_list,
                                                                        ordered=True)
                        bs_intensivnosti_full.sort_values(by='index', inplace=True)
                        bs_intensivnosti_full = bs_intensivnosti_full.reset_index(drop=True).set_index(
                            ['detector_id', 'index'])
                        bs_intensivnosti_to_excel = bs_intensivnosti_full.iloc[:,
                                                    bs_intensivnosti_full.columns.get_level_values(0) == direction]
                        bs_intensivnosti_to_excel.columns = bs_intensivnosti_to_excel.columns.droplevel(0)
                        bs_intensivnosti_to_excel = bs_intensivnosti_to_excel.reset_index(drop=False)
                        # append the dataframe results to the current excel file
                        for row in dataframe_to_rows(bs_intensivnosti_to_excel, header=False, index=False):
                            sheet_basic_stats.append(row)
                    except KeyError:  # если листа с таким названием нет
                        wb_basic_stats.create_sheet(direction)
                        sheet_basic_stats = wb_basic_stats[direction]  # declare the active sheet
                        bs_intensivnosti_full = bs_intensivnosti_full.reset_index().rename(
                            columns={'level_0': 'detector_id'})
                        bs_intensivnosti_full['index'] = pd.Categorical(bs_intensivnosti_full['index'],
                                                                        categories=month_list + weekdays_list + hours_list,
                                                                        ordered=True)
                        bs_intensivnosti_full.sort_values(by='index', inplace=True)
                        bs_intensivnosti_full = bs_intensivnosti_full.reset_index(drop=True).set_index(
                            ['detector_id', 'index'])
                        bs_intensivnosti_to_excel = bs_intensivnosti_full.iloc[:,
                                                    bs_intensivnosti_full.columns.get_level_values(0) == direction]
                        bs_intensivnosti_to_excel.columns = bs_intensivnosti_to_excel.columns.droplevel(0)
                        bs_intensivnosti_to_excel = bs_intensivnosti_to_excel.reset_index(drop=False)
                        # append the dataframe results to the current excel file
                        for row in dataframe_to_rows(bs_intensivnosti_to_excel, header=True, index=False):
                            sheet_basic_stats.append(row)
                wb_basic_stats.save(f'../out/{company_choice}/basic_stats_intensivnosti.xlsx')  # save workbook
                wb_basic_stats.close()  # close workbook
            else:  # create the excel file if doesn't already exist
                for direction in ['Итого', 'Прямое', 'Обратное']:
                    bs_intensivnosti_full = bs_intensivnosti_full.reset_index().rename(
                        columns={'level_0': 'detector_id'})
                    bs_intensivnosti_full['index'] = pd.Categorical(bs_intensivnosti_full['index'],
                                                                    categories=month_list + weekdays_list + hours_list,
                                                                    ordered=True)
                    bs_intensivnosti_full.sort_values(by='index', inplace=True)
                    bs_intensivnosti_full = bs_intensivnosti_full.reset_index(drop=True).set_index(
                        ['detector_id', 'index'])
                    bs_intensivnosti_to_excel = bs_intensivnosti_full.iloc[:,
                                                bs_intensivnosti_full.columns.get_level_values(0) == direction]
                    bs_intensivnosti_to_excel.columns = bs_intensivnosti_to_excel.columns.droplevel(0)
                    bs_intensivnosti_to_excel = bs_intensivnosti_to_excel.reset_index(drop=False)
                    # .assign(idx=pd.factorize(bs_intensivnosti_full['index'])[0]) \
                    # .pivot(index='detector_id', columns=['idx', 'index'], values=direction) \
                    # .droplevel('idx', axis=1).reset_index()
                    with pd.ExcelWriter(path=f'../out/{company_choice}/basic_stats_intensivnosti.xlsx', engine='openpyxl') as writer:
                        bs_intensivnosti_to_excel.to_excel(writer, index=False, header=True, sheet_name=direction)

            self.checking.plot_graphs(self.company.get(), df_main_clear, df_total_long, self.cur_year, new_file_name, freq='d')
        wb_data.app.quit()

    # def init_data_import(self, file, rbt_var=None):
    #     # Функция для вставки исходных данных от детектора в xls-шаблон для последующей обработки.
    #     # Копирует требуемые диапазоны ячеек в существующий xls-файл.
    #
    #     # Определяю номер последней строки в исходной таблице
    #     wb = openpyxl.load_workbook(os.path.join(self.directory_pre, file))
    #     sheet = wb.active
    #     columnA = sheet['A']
    #     last_row = len(sheet['A'])
    #     #    print(last_row)
    #
    #     # # Open the Excel program, the default setting: the program is visible, only open without creating a new workbook, and the screen update is turned off
    #     # app=xw.App(visible=True, add_book=False)
    #     # app.display_alerts=False
    #     # app.screen_updating=False
    #
    #     wb_data = xw.Book(os.path.join(self.directory_pre, file))
    #     wb_sample = None
    #
    #     type_file = None
    #     # Структуру исходных файлов не проверяю, поскольку сделал это уже ранее. Единый формат отчёта
    #     if rbt_var:
    #         type_file = self.structure_check(file)
    #
    #     # Проверка соответствия структуры исходных файлов отключена, поскольку сделал такую проверку ранее отдельно.
    #     # В текущей версии все файлы от Росавтодора - единого формата.
    #     if type_file:
    #         if type_file == 'Rosautodor_1':
    #             wb_sample = xw.Book('../sample_fda_var1.xlsm')  # 'sample_r1.xlsx'
    #         elif type_file == 'Rosautodor_2':
    #             wb_sample = xw.Book('../sample_fda_var2.xlsm')  # 'sample_r.xlsx'
    #     else:
    #         # if not type_file:
    #         tk.messagebox.showwarning(title="ALERT",
    #                                   message="Вы выбрали не проверять структура файла. Выберите файл шаблона вручную\n"
    #                                           "Выберите файл шаблона (sample_fda_var1.xlsm) или (sample_fda_var2.xlsm)\n"
    #                                           "Посмотреть какой именно шаблон подходит к обрабатываемому файлу можно посмотреть в log-файле в корне проекта")
    #         folder_path_to_sample = filedialog.askopenfile()  # Open a folder selection dialog
    #
    #         print(folder_path_to_sample)
    #
    #         if folder_path_to_sample:
    #             # print(True)
    #             wb_sample = xw.Book(folder_path_to_sample.name)
    #         elif not folder_path_to_sample:
    #             # print(False)
    #             tk.messagebox.showerror(title="ALERT",
    #                                     message="Не выбран ни один файл шаблона.\nВыберите и повторите попытку")
    #             print("Не выбран ни один файл. Выберите :", sys.exc_info()[0])
    #             return
    #
    #         # wb_sample = xw.Book('../sample_fda_var2.xlsm')
    #
    #     # Перевод окон в полноэкранный режим. Макрос прописан в файле шаблона 'sample_r1.xlsm'
    #     # Текст макроса для VBA (прописывается в Excel):
    #     # Sub Maximize_Window()
    #     # Application.WindowState = xlMaximized
    #     # End Sub
    #
    #     Maximize = wb_sample.macro('Maximize_Window')
    #     Maximize()
    #     time.sleep(5)
    #
    #     road_name = wb_data.sheets[0]['A6'].value
    #     road_name = road_name.replace('/', '.')  # Слеш заменяю на точку
    #     road_name = road_name.replace('"',
    #                                   '')  # Удаляю символы кавычек, вопросительный и восклицательный знаки, символ звёздочки, которые могут мешать сохранению файла
    #     road_name = road_name.replace('?', '')
    #     road_name = road_name.replace('!', '')
    #     road_name = road_name.replace('*', '')
    #     road_name = road_name.replace('\n', '')  # Убираю символы переноса строки
    #     road_name = road_name.replace('\t', '')  # Убираю символы табуляции
    #
    #     # Копирую данные о названии дороги и местоположении детектора
    #     my_values = wb_data.sheets[0].range('A5:A6').options(ndim=2).value
    #     wb_sample.sheets['Исходные данные'].range('A3:A4').value = my_values
    #
    #     # Сведения о категориях ТС не копируются, поскольку уже заложены в шаблон 'sample_r1.xlsx'
    #
    #     # Изначально копировал исходные данные об интенсивности движения только по 24*366=8784 строкам
    #     # Выяснилось, что в отчётах могут быть данные за несколько лет и копировать нужно все данные
    #     my_values = wb_data.sheets[0].range('A7:AB' + str(last_row)).options(ndim=2).value
    #     wb_sample.sheets['Исходные данные'].range('A5:AB' + str(last_row - 2)).value = my_values
    #
    #     # wb_sample.sheets['Итоги'].activate()
    #     # time.sleep(5)
    #     # image = pyscreenshot.grab()
    #     # # print(file)
    #     # image.save(os.path.join('Графики', filename[0:len(filename) - 5] + '.png'))
    #
    #     # # Копирую заголовок с описанием типов транспортных средств
    #     # car_category=wb_data.sheets['Исходные данные'].range('E7:AB7').options(ndim=2).value
    #     # wb_sample.sheets['Исходные данные'].range('E1:AB1').value = car_category
    #
    #     #     # Перехожу на вкладку "Итоги" и после этого ожидаю 5 секунд для отрисовки графиков
    #     # #    wb_sample.sheets["Итоги"].api.Tab.Color = rgb_to_int((146, 208, 80))
    #     #     wb_sample.sheets['Итоги'].activate()
    #     #     print('OK')
    #     #     time.sleep(5)
    #
    #     # Сохранение под новым названием в специальной папке
    #     # wb_sample.save(os.path.join('Импортированные данные', 'IN_'+road_name+'.xlsx'))
    #     # df_long = os.path.join(directory_imp, 'IN_' + road_name + file[4:])
    #     wb_sample.save(os.path.join(self.directory_imp, 'IN_' + file[4:]))
    #
    #     # делаю скриншоты сразу
    #     # wb_data = xw.Book(os.path.join(directory_imp, file))
    #     wb_sample.sheets['Итоги'].activate()
    #     # Maximize = wb_data.macro('Maximize_Window')
    #     # Maximize()
    #     time.sleep(3)
    #     image = pyscreenshot.grab()
    #     # print(file)
    #     # short_filename = re.search('IN_.+', str(file))[0][3:]
    #     # Первое совпадение, начинающееся на "флаг" IN_, с отбрасыванием этого флага при формировании имени файла (т.е. в имя файла включаются символы, начиная с четвёртого)
    #     # print(short_filename)
    #     # image.save(os.path.join('Графики', file[0:len(short_filename) - 5] + '.png'))
    #     image.save(os.path.join(self.directory_pic, 'PIC_' + file[3:len(file) - 5] + '.png'))
    #     wb_data.app.quit()
    #     # wb_sample.app.quit()
    #
    # def screenshots(self, file):
    #     wb_service = xw.Book('../maximize.xlsm')
    #     Maximize = wb_service.macro('Maximize_Window')
    #     Maximize()
    #
    #     wb_data = xw.Book(os.path.join(self.directory_imp, file))
    #     wb_data.sheets['Итоги'].activate()
    #     # Maximize = wb_data.macro('Maximize_Window')
    #     # Maximize()
    #     time.sleep(3)
    #     image = pyscreenshot.grab()
    #     # print(file)
    #     # short_filename = re.search('IN_.+', str(file))[0][3:]
    #     # Первое совпадение, начинающееся на "флаг" IN_, с отбрасыванием этого флага при формировании имени файла (т.е. в имя файла включаются символы, начиная с четвёртого)
    #     # print(short_filename)
    #     # image.save(os.path.join('Графики', file[0:len(short_filename) - 5] + '.png'))
    #     image.save(os.path.join(self.directory_pic, 'PIC_' + file[3:len(file) - 5] + '.png'))
    #     wb_data.app.quit()


# if __name__ == "__main__":
#     root = tk.Tk()  # окно
#     app = Application(root)  # ПУСК
#     root.mainloop()
