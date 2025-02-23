import time
import numpy as np
import pandas as pd
import openpyxl
import re
from re import search
import os
import sys
import xlwings as xw
from xlwings.utils import rgb_to_int, int_to_rgb
import logging
import pyscreenshot
from tqdm import tqdm
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter.ttk import Combobox
from tkinter.scrolledtext import ScrolledText
from tkinter import messagebox

cur_year = '2024'
directory_raw = f'../raw_data/{cur_year}/Исходные данные'
directory_pre = f'../raw_data/{cur_year}/Первичная обработка'
directory_imp = f'../raw_data/{cur_year}/Импортированные данные'
directory_pic = f'../raw_data/{cur_year}/Графики'

for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(message)s',
                    datefmt='%a, %d %b %Y %H:%M:%S',
                    filename='log.txt',
                    filemode='w')

window = Tk()
window.title("Добро пожаловать в приложение! ФДА РОСАВТОДОР")
window.geometry("1200x620")

lbl = Label(window, text="\tВыберите директорию с исходными файлами или папками с файлами\n"
                         "\tДля обратки только что поступивших файлов датчиков можно выбрать папку 'Исходные данные', даже если там находятся подпапки и далее выделить нужные файлы/папки и нажать запустить\n"
                         "\tДля обработки файлов, с проверенной структурой можно выбрать папку 'Первичная обработка' (с префиксом PRE_)\n"
                         "\tДля отрисовки графиков выбрать папку 'Импортированные данные' (с префиксом IN_)",
            anchor="w", justify=LEFT)
lbl.grid(row=0, columnspan=7)  # .pack(pady=10)

folder_path = None


# Function to update the Listbox with folder contents
def browse_folder():
    global folder_path
    folder_path = filedialog.askdirectory()  # Open a folder selection dialog
    if folder_path:
        folder_contents.delete(0, tk.END)  # Clear the Listbox
        for item in os.listdir(folder_path):
            folder_contents.insert(tk.END, item)  # Insert folder contents into Listbox


# Create a button to browse for a folder
browse_button = tk.Button(window, text="Browse Folder", command=browse_folder)
browse_button.grid(row=1, column=0, sticky=W, pady=10, padx=10)  # .pack(pady=10)

# Create a Listbox to display folder contents
folder_contents = tk.Listbox(window, selectmode=tk.EXTENDED, exportselection=False)
folder_contents.grid(row=2, column=0, columnspan=5, sticky=NSEW, padx=10)  # .pack(fill=tk.BOTH, expand=True)

# Create a scrollbar for the Listbox
scrollbar = tk.Scrollbar(folder_contents)
scrollbar.pack(side=RIGHT, fill=Y)
folder_contents.configure(yscrollcommand=scrollbar.set)
scrollbar.configure(command=folder_contents.yview)


def select_all():
    folder_contents.select_set(0, tk.END)


select_button = tk.Button(window, text="Select All", command=select_all)
select_button.grid(row=3, column=2, sticky=W, pady=10, padx=10)  # .pack(pady=10)
# select_button.grid(row=5, column=2)

# Create a textbox to display log and other information

# editor = tk.Text(window, wrap='word', width=40, height=15)
# editor.grid(row=2, column=6, padx=10, sticky=NSEW)
#
# ys = tk.Scrollbar(orient="vertical", command=editor.yview)
# # ys.grid(column=1, row=0, sticky=NS)
# xs = tk.Scrollbar(orient="horizontal", command=editor.xview)
# # xs.grid(column=0, row=1, sticky=EW)
#
# editor["yscrollcommand"] = ys.set
# editor["xscrollcommand"] = xs.set

editor = ScrolledText(window, wrap='word')
editor.grid(row=2, column=6, padx=(10, 10), sticky=NE)


def structure_check(file):
    # print(f'\n\n====== FILENAME FROM FUNCTION {os.path.join(directory_raw, file)}\n\n')
    # file = 'PRE_км 41+138 а.д М-10 Россия Москва – Санкт-Петербург .xlsx'
    filename_prefix = re.match('^PRE_.+', str(file))
    if filename_prefix:
        wb = openpyxl.load_workbook(os.path.join(directory_pre, file))
    else:
        wb = openpyxl.load_workbook(os.path.join(directory_raw, file))
    sheet = wb.active

    # 1. Проверка соответствия отчёта заданной форме (с записью итогов проверки в лог-файл)

    if sheet['B3'].value == "Общая интенсивность автомобилей" and sheet['E3'].value == "Легковые (до 6 м)" and sheet[
        'H3'].value == "Малые груз. (6-9 м)" and sheet['K3'].value == "Грузовые (9-13 м)" and \
            sheet['N3'].value == "Груз. большие (13-22 м)" and sheet['Q3'].value == "Автопоезда (22-30 м)" and sheet[
        'T3'].value == "Автобусы" and sheet['W3'].value == "Мотоциклы":
        logging.info(f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №1')
        editor.insert(tk.END, f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №1\nИспользую шаблон отчета (pre_sample_r1.xlsx)\n')
        return 'Rosautodor_1'
    elif sheet['B3'].value == "Общая интенсивность автомобилей" and sheet['E3'].value == "Легковые (до 4.5 м)" and \
            sheet[
                'H3'].value == "Легковые большие (4-6 м)" and sheet[
        'K3'].value == "Малые груз. (6-9 м)" and sheet['N3'].value == "Грузовые (9-13 м)" and \
            sheet['Q3'].value == "Груз. большие (13-22 м)" and sheet['T3'].value == "Автопоезда (22-30 м)" and sheet[
        'W3'].value == "Автобусы" and sheet['Z3'].value == "Мотоциклы":
        logging.info(f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №2')
        editor.insert(tk.END, f'{file}: Формат отчёта соответствует 7 категориям ТС, Варианту №2\nИспользую шаблон отчета (pre_sample_r2.xlsx)\n')
        return 'Rosautodor_2'

    # elif sheet['B6'].value=="За период" and sheet['E7'].value=="легковые автомобили (до 6 м)" and sheet['H7'].value=="микроавтобусы, малые грузовые автомобили (6-9 м)" and sheet['K7'].value=="грузовые автомобили (9-11 м)" and\
    #         sheet['N7'].value=="автобусы (11-13 м)" and sheet['Q7'].value=="грузовые большие автомобили, автопоезда (13-18 м)" and sheet['T7'].value=="длинные автопоезда (> 18 м)":
    #     logging.info(f'{filename}: Формат отчёта соответствует 6 категориям ТС, Варианту №2')
    # elif sheet['B6'].value == "За период" and sheet['E7'].value == "легковые автомобили (до 6 м)" and sheet['H7'].value == "малые грузовые автомобили до 5 тонн (6-9 м)" and\
    #         sheet['K7'].value == "грузовые автомобили 5-12 тонн (9-11 м)" and sheet['N7'].value == "автобусы (11-13 м)" and sheet['Q7'].value == "грузовые большие автомобили 12-20 тонн (13-22 м)" and\
    #         sheet['T7'].value == "автопоезда более 20 тонн (22-30 м)":
    #     logging.info(f'{filename}: Формат отчёта соответствует 6 категориям ТС, Варианту №3')
    #     editor.insert(tk.END, f'{file}: Формат отчёта соответствует 6 категориям ТС, Варианту №3')
    else:
        logging.info(f'{file}: Формат отчёта не совпадает')
        editor.insert(tk.END, f'{file}: Формат отчёта не совпадает')
        return 0


def preprocessing(file, rbt_var=None):

    wb = openpyxl.load_workbook(os.path.join(directory_raw, file))  # открываем файл с данными
    sheet = wb.active  # активируем лист
    columnA = sheet['A']  # запоминаем колонку А

    dataset_begins = []  # список ячеек, где хранятся тектовые данные для аккуратных наименований

    for cell in columnA:
        if (search("км", str(cell.value)) or search("km", str(cell.value))) and cell.row > 4:
            dataset_begins = dataset_begins + [cell.row]
    dataset_begins = dataset_begins + [len(columnA) + 1]

    type_file = None
    # Старый коммента - # Структуру исходных файлов не проверяю, поскольку сделал это уже ранее. Единый формат отчёта
    # Теперь структуру проверяем (можно выбрать проверять или нет с помощью RadioButton)
    if rbt_var:  # если RadioButton выбран "да", то происходит проверка структуры и в type_file возвращается тип структуры 'Rosautodor_1' или 'Rosautodor_2'
        type_file = structure_check(file)


    wb_data = xw.Book(os.path.join(directory_raw, file))

    # Копирую данные о названии дороги и местоположении детектора, категориях ТС
    for i in range(0, len(dataset_begins) - 1):
        # ниже старые комменты
        # Проверка соответствия структуры исходных файлов отключена, поскольку сделал такую проверку ранее отдельно.
        # В текущей версии все файлы от Росавтодора - единого формата.

        # теперь если включена проверка структуры, то в зависимости от того, что вернулось, тот шаблон и открываем и вставляем туда данные
        if type_file:
            if type_file == 'Rosautodor_1':
                wb_sample = xw.Book('../pre_sample_r1.xlsx')  # pre_sample_r1.xlsx
            elif type_file == 'Rosautodor_2':
                wb_sample = xw.Book('../pre_sample_r2.xlsx')  # pre_sample_r2.xlsx
        else:
            wb_sample = xw.Book('../pre_sample_r2.xlsx')  # pre_sample_r2.xlsx

        road = wb_data.sheets[0]['A5'].value
        wb_sample.sheets['Исходные данные']['A5'].value = road

        place = wb_data.sheets[0]['A' + str(dataset_begins[i])].value
        wb_sample.sheets['Исходные данные']['A6'].value = place

        place = place.replace('/', '.')  # Слеш заменяю на точку
        place = place.replace('"', '')  # Удаляю символы кавычек, вопросительный и восклицательный знаки, символ звёздочки, которые могут мешать сохранению файла
        place = place.replace('?', '')
        place = place.replace('!', '')
        place = place.replace('*', '')
        place = place.replace('\n', '')  # Убираю символы переноса строки
        place = place.replace('\t', '')  # Убираю символы табуляции

        # Копирую исходные данные об интенсивности движения
        my_values = wb_data.sheets[0].range(
            str('A' + str(dataset_begins[i] + 1) + ':AB' + str(dataset_begins[i + 1] - 1))).options(ndim=2).value
        wb_sample.sheets['Исходные данные'].range('A7').value = my_values

        # # Копирую заголовок с описанием типов транспортных средств
        # car_category=wb_data.sheets['Исходные данные'].range('E7:AB7').options(ndim=2).value
        # wb_sample.sheets['Исходные данные'].range('E1:AB1').value = car_category

        # Сохранение под новым названием в специальной папке
        # wb_sample.save(os.path.join('Первичная обработка', 'PRE_' + place + '.xlsx'))
        wb_sample.save(os.path.join(directory_pre, 'PRE_' + place + '.xlsx'))
        # wb_sample.close()

    wb_data.app.quit()


def init_data_import(file, rbt_var=None):
    # Функция для вставки исходных данных от детектора в xls-шаблон для последующей обработки.
    # Копирует требуемые диапазоны ячеек в существующий xls-файл.

    # Определяю номер последней строки в исходной таблице
    wb = openpyxl.load_workbook(os.path.join(directory_pre, file))
    sheet = wb.active
    columnA = sheet['A']
    last_row = len(sheet['A'])
    #    print(last_row)

    # # Open the Excel program, the default setting: the program is visible, only open without creating a new workbook, and the screen update is turned off
    # app=xw.App(visible=True, add_book=False)
    # app.display_alerts=False
    # app.screen_updating=False

    wb_data = xw.Book(os.path.join(directory_pre, file))

    type_file = None
    # Структуру исходных файлов не проверяю, поскольку сделал это уже ранее. Единый формат отчёта
    if rbt_var:
        type_file = structure_check(file)

    # Проверка соответствия структуры исходных файлов отключена, поскольку сделал такую проверку ранее отдельно.
    # В текущей версии все файлы от Росавтодора - единого формата.
    if type_file:
        if type_file == 'Rosautodor_1':
            wb_sample = xw.Book('../sample_fda_var1.xlsm')  # 'sample_r1.xlsx'
        elif type_file == 'Rosautodor_2':
            wb_sample = xw.Book('../sample_fda_var2.xlsm')  # 'sample_r.xlsx'
    else:
        wb_sample = xw.Book('../sample_fda_var2.xlsm')

    # Перевод окон в полноэкранный режим. Макрос прописан в файле шаблона 'sample_r1.xlsm'
    # Текст макроса для VBA (прописывается в Excel):
    # Sub Maximize_Window()
    # Application.WindowState = xlMaximized
    # End Sub

    Maximize = wb_sample.macro('Maximize_Window')
    Maximize()
    time.sleep(5)

    road_name = wb_data.sheets[0]['A6'].value
    road_name = road_name.replace('/', '.')  # Слеш заменяю на точку
    road_name = road_name.replace('"',
                                  '')  # Удаляю символы кавычек, вопросительный и восклицательный знаки, символ звёздочки, которые могут мешать сохранению файла
    road_name = road_name.replace('?', '')
    road_name = road_name.replace('!', '')
    road_name = road_name.replace('*', '')
    road_name = road_name.replace('\n', '')  # Убираю символы переноса строки
    road_name = road_name.replace('\t', '')  # Убираю символы табуляции

    # Копирую данные о названии дороги и местоположении детектора
    my_values = wb_data.sheets[0].range('A5:A6').options(ndim=2).value
    wb_sample.sheets['Исходные данные'].range('A3:A4').value = my_values

    # Сведения о категориях ТС не копируются, поскольку уже заложены в шаблон 'sample_r1.xlsx'

    # Изначально копировал исходные данные об интенсивности движения только по 24*366=8784 строкам
    # Выяснилось, что в отчётах могут быть данные за несколько лет и копировать нужно все данные
    my_values = wb_data.sheets[0].range('A7:AB' + str(last_row)).options(ndim=2).value
    wb_sample.sheets['Исходные данные'].range('A5:AB' + str(last_row - 2)).value = my_values

    # wb_sample.sheets['Итоги'].activate()
    # time.sleep(5)
    # image = pyscreenshot.grab()
    # # print(file)
    # image.save(os.path.join('Графики', filename[0:len(filename) - 5] + '.png'))

    # # Копирую заголовок с описанием типов транспортных средств
    # car_category=wb_data.sheets['Исходные данные'].range('E7:AB7').options(ndim=2).value
    # wb_sample.sheets['Исходные данные'].range('E1:AB1').value = car_category

    #     # Перехожу на вкладку "Итоги" и после этого ожидаю 5 секунд для отрисовки графиков
    # #    wb_sample.sheets["Итоги"].api.Tab.Color = rgb_to_int((146, 208, 80))
    #     wb_sample.sheets['Итоги'].activate()
    #     print('OK')
    #     time.sleep(5)

    # Сохранение под новым названием в специальной папке
    # wb_sample.save(os.path.join('Импортированные данные', 'IN_'+road_name+'.xlsx'))
    # df_long = os.path.join(directory_imp, 'IN_' + road_name + file[4:])
    wb_sample.save(os.path.join(directory_imp, 'IN_' + file[4:]))

    # делаю скриншоты сразу
    # wb_data = xw.Book(os.path.join(directory_imp, file))
    wb_sample.sheets['Итоги'].activate()
    # Maximize = wb_data.macro('Maximize_Window')
    # Maximize()
    time.sleep(3)
    image = pyscreenshot.grab()
    # print(file)
    # short_filename = re.search('IN_.+', str(file))[0][3:]
    # Первое совпадение, начинающееся на "флаг" IN_, с отбрасыванием этого флага при формировании имени файла (т.е. в имя файла включаются символы, начиная с четвёртого)
    # print(short_filename)
    # image.save(os.path.join('Графики', file[0:len(short_filename) - 5] + '.png'))
    image.save(os.path.join(directory_pic, 'PIC_' + file[3:len(file) - 5] + '.png'))
    wb_data.app.quit()
    # wb_sample.app.quit()


# def screenshots(file):
#     wb_service = xw.Book('../maximize.xlsm')
#     Maximize = wb_service.macro('Maximize_Window')
#     Maximize()
#
#     wb_data = xw.Book(os.path.join(directory_imp, file))
#     wb_data.sheets['Итоги'].activate()
#     # Maximize = wb_data.macro('Maximize_Window')
#     # Maximize()
#     time.sleep(3)
#     image = pyscreenshot.grab()
#     # print(file)
#     # short_filename = re.search('IN_.+', str(file))[0][3:]
#     # Первое совпадение, начинающееся на "флаг" IN_, с отбрасыванием этого флага при формировании имени файла (т.е. в имя файла включаются символы, начиная с четвёртого)
#     # print(short_filename)
#     # image.save(os.path.join('Графики', file[0:len(short_filename) - 5] + '.png'))
#     image.save(os.path.join(directory_pic, 'PIC_' + file[3:len(file) - 5] + '.png'))
#     wb_data.app.quit()


def selected(event, rbt_var):
    selection = combobox.get()
    tmp = [folder_contents.get(idx) for idx in
           folder_contents.curselection()]  # список папок или файлов из Combobox в окошке
    if not tmp:
        tk.messagebox.showerror(title="ALERT", message="Не выбран ни один файл или папка из списка.\nВыберите и повторите попытку")
        return
    # print(tmp)
    xlsx_files = []  # список файлов, если в папке "Исходные данные" не файлы, а аще папки
    for i in tmp:  # для каждого файла или папки из Combobox проверяем файл это или папка
        if os.path.isfile(os.path.join(folder_path, i)):
            xlsx_files.append(i)  # если файл, то все ок - оставляем список как есть
        else:
            for file in os.listdir(os.path.join(folder_path, i)):
                xlsx_files.append(os.path.join(i,
                                               file))  # если папка, не ок - открываем каждую папку и смотрим что там за файл и добавляем в новый список

    if selection == "Проверка структуры файлов":
        editor.delete(tk.END)
        print(f'Выбор 1.\nПроверяю файлы {chr(10).join(xlsx_files)} на соответствие структуре\n')
        editor.insert(tk.END, f'\nВыбор 1.\nПроверяю файлы {chr(10).join(xlsx_files)} на соответствие структуре\n')
        for filename in tqdm(xlsx_files):
            editor.insert(tk.END, '\n======' + filename + '======\n')
            f = os.path.join(folder_path, filename)
            if os.path.isfile(f):
                window.update_idletasks()
                time.sleep(2)
                structure_check(f)
    elif selection == "Предобработка":
        print(f'Выбор 2.\nПредобрабатываю файлы {chr(10).join(xlsx_files)}...\n')
        # combobox.bind("<<ComboboxSelected>>", lambda: showDeleteOptions())
        # lbl.grid(row=4, column=0, padx=10)  # .pack(pady=10)
        # yes_rdb.grid(row=5, column=0, padx=10, sticky=W)
        # no_rdb.grid(row=6, column=0, padx=10, sticky=W)
        editor.delete(tk.END)
        window.update_idletasks()
        editor.insert(tk.END, f'\nВыбор 2.\nПредобрабатываю файлы {chr(10).join(xlsx_files)}...\n')
        for filename in tqdm(xlsx_files):
            editor.see(END)
            editor.insert(tk.END, f'\n======' + filename + '======\n')
            f = os.path.join(folder_path, filename)
            # checking if it is a file
            if os.path.isfile(f):
                window.update_idletasks()
                time.sleep(1)
                if rbt_var == 'yes':
                    editor.insert(tk.END, f'Включена проверка файла на соотвествтие структуре...\n'
                                          f'Проверяю сначала структуру файла {filename}\n')
                    preprocessing(filename, var)
                elif rbt_var == 'no':
                    preprocessing(filename)
            editor.insert(tk.END, f'\n{filename} обработан!\n')
    elif selection == "Вставка обработанных данных в шаблон .xlsm":
        combobox.bind("<<ComboboxSelected>>", lambda: showDeleteOptions())
        print(f'\nВыбор 3.\nВставляю данные файлов {chr(10).join(xlsx_files)} в шаблон отчета...\n')
        # lbl.grid(row=4, column=0, padx=10)  # .pack(pady=10)
        # yes_rdb.grid(row=5, column=0, padx=10, sticky=W)
        # no_rdb.grid(row=6, column=0, padx=10, sticky=W)
        editor.delete(tk.END)
        editor.insert(tk.END, f'\nВыбор 3.\nВставляю данные файлов {chr(10).join(xlsx_files)} в шаблон отчета...\n')
        for filename in tqdm(xlsx_files):
            editor.insert(tk.END, '\n======' + filename + '======\n')
            editor.see(END)
            f = os.path.join(folder_path, filename)
            # checking if it is a file
            if os.path.isfile(f):
                window.update_idletasks()
                time.sleep(2)
                if rbt_var == 'yes':
                    editor.insert(tk.END, f'Включена проверка файла на соотвествтие структуре...\n'
                                          f'Проверяю сначала структуру файла {filename}\n')
                    init_data_import(filename, var)
                elif rbt_var == 'no':
                    init_data_import(filename, var)
            editor.insert(tk.END, f'Файл {filename} обработан!\n')
            editor.see(END)
    # elif selection == "Скриншоты":
    #     print('\nВыбор 4.\nСоздаю скриншоты\n')
    #     editor.delete(tk.END)
    #     tmp = [folder_contents.get(idx) for idx in folder_contents.curselection()]
    #     editor.insert(tk.END, f'\nВыбор 4.\nСоздаю скриншоты файлов {chr(10).join(tmp)}...\n')
    #     for filename in tqdm(tmp):
    #         editor.insert(tk.END, '\n======' + filename + '======\n')
    #         f = os.path.join(folder_path, filename)
    #         # checking if it is a file
    #         if os.path.isfile(f):
    #             window.update_idletasks()
    #             time.sleep(2)
    #             screenshots(filename)
    #         editor.insert(tk.END, f'{filename} обработан!\n')


def DeleteOptions(event):
    if combobox.get() == "Проверка структуры файлов":
        hideDeleteOptions()
    elif combobox.get() == 'Предобработка' or combobox.get() == 'Вставка обработанных данных в шаблон .xlsm':
        showDeleteOptions()

def hideDeleteOptions():
    lbl.grid_forget()
    yes_rdb.grid_forget()
    no_rdb.grid_forget()


def showDeleteOptions():
    lbl.grid(row=4, column=0, padx=10)  # .pack(pady=10)
    yes_rdb.grid(row=5, column=0, padx=10, sticky=W)
    no_rdb.grid(row=6, column=0, padx=10, sticky=W)


choice = ['Проверка структуры файлов', 'Предобработка', 'Вставка обработанных данных в шаблон .xlsm']  # , 'Скриншоты'
combobox = Combobox(window, values=choice, width=30, state="readonly")
combobox.grid(row=3, column=0, pady=10, padx=10)
combobox.bind("<<ComboboxSelected>>", DeleteOptions)

lbl = Label(window, text="Проверять тип файла перед обработкой?")

var = StringVar()
var.set("yes")
yes_rdb = Radiobutton(window, text='Да', variable=var, value="yes")
no_rdb = Radiobutton(window, text='Нет', variable=var, value="no")

run_btn = Button(
    window,
    text="Запустить",
    command=lambda: selected(combobox.get(), var.get())
)
run_btn.grid(row=5, column=2, padx=10, pady=10)

window.mainloop()

